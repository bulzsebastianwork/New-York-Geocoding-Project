from openpyxl import load_workbook
import os


wb_metros = load_workbook("D:\\Upwork Projects\\Python\\NYC Metro Stations.xlsx")
ws_metros = wb_metros.active
for filename in os.listdir("D:\\Upwork Projects\\Python\\1. Amit's Metro Coordinates to Table"):
    if "brooklyn" in filename:
        wb = load_workbook(os.path.join("D:\\Upwork Projects\\Python\\1. Amit's Metro Coordinates to Table", filename))
        ws = wb.active

        counter_rows = 6
        while not(ws.cell(counter_rows, 9).value is None):
            station_name = str(ws.cell(counter_rows, 25).value)

            counter_rows_metros = 2
            while station_name != str(ws_metros.cell(counter_rows_metros, 1).value):
                counter_rows_metros += 1

            split_coordinates = str(ws_metros.cell(counter_rows_metros, 3).value).split(" ")
            ws.cell(counter_rows, 27).value = split_coordinates[1] + ", " + split_coordinates[0]
            print("Writing Coordinates for " + filename[:-4] + " Counter: " + str(counter_rows))

            counter_rows += 1

        wb.save(filename.replace(".xlsx", "_coordinates.xlsx"))
        wb.close()