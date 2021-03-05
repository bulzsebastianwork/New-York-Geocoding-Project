from openpyxl import load_workbook
from geopy.distance import geodesic
import time

def main():

    wb_addresses = load_workbook("manhattan_found_2019.xlsx")
    ws_addresses = wb_addresses.active

    wb_metros = load_workbook("NYC Metro Stations.xlsx")
    ws_metros = wb_metros.active

    Counter_Rows = 6

    while not(ws_addresses.cell(Counter_Rows, 11).value is None):
        Counter_Rows_2 = 2

        min_distance = 100
        while not(ws_metros.cell(Counter_Rows_2, 3).value is None):
            split_coordinates_comma = str(ws_metros.cell(Counter_Rows_2, 3).value).replace(",", "")
            split_coordinates_metro = str(split_coordinates_comma).split(" ")
            metro_tuple = (float(split_coordinates_metro[1]), float(split_coordinates_metro[0]))

            split_coordinates_table = str(ws_addresses.cell(Counter_Rows, 11).value).replace(",", "").split(" ")

            table_tuple = (float(split_coordinates_table[0]), float(split_coordinates_table[1]))

            if geodesic(metro_tuple, table_tuple).miles < min_distance:
                min_distance = geodesic(metro_tuple, table_tuple).miles
                station_name_saved = str(ws_metros.cell(Counter_Rows_2, 1).value)
                station_address_saved = str(ws_metros.cell(Counter_Rows_2, 2).value)
                coordinates_split = str(ws_metros.cell(Counter_Rows_2, 3).value).split(" ")
                coordinates_saved = coordinates_split[1] + ", " + coordinates_split[0]

            Counter_Rows_2 += 1

        print (str(Counter_Rows) + ". " + "Dist=" + str(min_distance) + " " + station_name_saved + " " + station_address_saved)
        ws_addresses.cell(Counter_Rows, 24).value = min_distance
        ws_addresses.cell(Counter_Rows, 25).value = station_name_saved
        ws_addresses.cell(Counter_Rows, 26).value = station_address_saved
        ws_addresses.cell(Counter_Rows, 27).value = coordinates_saved
  
        Counter_Rows += 1

    wb_addresses.save("manhattan_found_metros_2019.xlsx")
    wb_addresses.close()

main()