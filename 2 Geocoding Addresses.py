import difflib
from geopy import Nominatim
from openpyxl import load_workbook

def replace_geo_abbreviations(street):
    switcher = {
        "alley": "aly",
        "annex": "anx",
        "apartment": "apt",
        "arcade": "arc",
        "avenue": "ave",
        "basement": "bsmt",
        "bayou": "byu",
        "beach": "bch",
        "bend": "bnd",
        "bluff": "blf",
        "bottom": "btm",
        "boulevard": "blvd",
        "branch": "br",
        "brook": "brookbrk",
        "building": "bldg",
        "burg": "bg",
        "bypass": "byp",
        "camp": "cp",
        "canyon": "cyn",
        "cape": "cpe",
        "causeway": "cswy",
        "center": "ctr",
        "circle": "cir",
        "cliff": "clfs",
        "cliffs": "clfs",
        "club": "clb",
        "corner": "cor",
        "corners": "cors",
        "course": "crse",
        "court": "ct",
        "courts": "cts",
        "cove": "cv",
        "creek": "crk",
        "crescent": "cres",
        "crossing": "xing",
        "dale": "dl",
        "dam": "dm",
        "department": "dept",
        "divide": "dv",
        "drive": "dr",
        "estate": "est",
        "expressway": "expy",
        "extension": "ext",
        "falls": "fls",
        "ferry": "fry",
        "field": "fld",
        "fields": "flds",
        "flat": "flt",
        "floor": "fl",
        "ford": "frd",
        "forest": "frst",
        "forge": "frg",
        "fork": "frk",
        "forks": "frks",
        "fort": "ft",
        "freeway": "fwy",
        "front": "frnt",
        "garden": "gdns",
        "gardens": "gdns",
        "gateway": "gtwy",
        "glen": "gln",
        "green": "grn",
        "grove": "grv",
        "hanger": "hngr",
        "harbor": "hbr",
        "haven": "hvn",
        "heights": "hts",
        "highway": "hwy",
        "hill": "hl",
        "hills": "hls",
        "hollow": "holw",
        "inlet": "inlt",
        "island": "is",
        "islands": "iss",
        "junction": "jct",
        "key": "ky",
        "knoll": "knls",
        "knolls": "knls",
        "lake": "lk",
        "lakes": "lks",
        "landing": "lndg",
        "lane": "ln",
        "light": "lgt",
        "loaf": "lf",
        "lobby": "lbby",
        "lock": "lcks",
        "locks": "lcks",
        "lodge": "ldg",
        "lower": "lowr",
        "manor": "mnr",
        "meadow": "mdws",
        "meadows": "mdws",
        "mill": "ml",
        "mills": "mls",
        "mission": "msn",
        "mount": "mt",
        "mountain": "mtn",
        "neck": "nck",
        "office": "ofc",
        "orchard": "orch",
        "parkway": "pkwy",
        "penthouse": "ph",
        "pine": "pnes",
        "pines": "pnes",
        "place": "pl",
        "plain": "pln",
        "plains": "plns",
        "plaza": "plz",
        "point": "pt",
        "port": "prt",
        "prairie": "pr",
        "radial": "radl",
        "ranch": "rnch",
        "rapid": "rpds",
        "rapids": "rpds",
        "rest": "rst",
        "ridge": "rdg",
        "river": "riv",
        "road": "rd",
        "room": "rm",
        "shoal": "shl",
        "shoals": "shls",
        "shore": "shr",
        "shores": "shrs",
        "space": "spc",
        "spring": "spg",
        "springs": "spgs",
        "square": "sq",
        "station": "sta",
        "stravenue": "stra",
        "stream": "strm",
        "street": "st",
        "suite": "ste",
        "summit": "smt",
        "terrace": "ter",
        "trace": "trce",
        "track": "trak",
        "trafficway": "trfy",
        "trail": "trl",
        "trailer": "trlr",
        "tunnel": "tunl",
        "turnpike": "tpke",
        "union": "un",
        "upper": "uppr",
        "valley": "vly",
        "viaduct": "via",
        "view": "vw",
        "village": "vlg",
        "ville": "vl",
        "vista": "vis",
        "well": "wls",
        "wells": "wls"
    }
    return switcher.get(street)

def search_for_correct_street(address, array_of_streets_brooklyn):

    counter_to_space = 0
    while address[counter_to_space] != " ":
        counter_to_space += 1

    while address[counter_to_space].isalpha() and counter_to_space < len(address) - 1:
        counter_to_space += 1

    if counter_to_space + 1 < len(address[counter_to_space]):
        counter_to_space -= 1
        if address[counter_to_space].isalpha() and address[counter_to_space + 1].isnumeric():
            address = address[:counter_to_space + 1] + " " + address[counter_to_space + 1:]

    address_split_space = address.split(" ")
    street_name = ""
    for counter_2 in range(1, len(address_split_space) - 1):
        street_name = street_name + address_split_space[counter_2] + " "

    try:
        street_name = street_name + replace_geo_abbreviations(address_split_space[counter_2 + 1].lower()).upper()
    except:
        street_name = street_name + address_split_space[len(address_split_space) - 1]

    address = address_split_space[0] + " " + street_name

    address_split_space = address.split(" ")
    street = ""
    for j in range(1, len(address_split_space) - 1):
        street = street + address_split_space[j] + " "

    street = street + address_split_space[len(address_split_space) - 1]
    street = street.upper()

    if len(difflib.get_close_matches(street, array_of_streets_brooklyn)) > 0:
        correct_street_name = difflib.get_close_matches(street, array_of_streets_brooklyn)
        if street[0] != correct_street_name[0][0]:
            correct_street_name = [street]
    else:
        correct_street_name = [street]

    address = address.replace(street, correct_street_name[0])

    return address

def main():
    streets = open("Manhattan_Streets.txt", "r")

    array_of_streets = streets.readlines()
    array_of_streets_brooklyn = list(dict.fromkeys(array_of_streets))

    for i in range(len(array_of_streets_brooklyn)):
        array_of_streets_brooklyn[i] = array_of_streets_brooklyn[i].upper()

    for i in range(len(array_of_streets_brooklyn)):
        array_of_streets_brooklyn[i] = array_of_streets_brooklyn[i].replace("\n", "")

    wb = load_workbook("manhattan_hash_2019.xlsx")
    ws = wb.active

    counter_rows = 5

    geolocator = Nominatim(user_agent = "geocoding_manhattan_2019")

    try:
        while not(ws.cell(counter_rows, 9).value is None):
            if ws.cell(counter_rows, 11).value is None and ws.cell(counter_rows, 21).value > 49999 and len(str(ws.cell(counter_rows, 9).value).split(" ")) > 0:
                if "," in str(ws.cell(counter_rows, 9).value):
                    address_without_comma = str(ws.cell(counter_rows, 9).value)
                    counter_comma = len(ws.cell(counter_rows, 9).value) - 1
                    while not("," in address_without_comma[counter_comma]):
                        counter_comma -= 1

                    address_without_comma = address_without_comma[:counter_comma - 1]
                else:
                    address_without_comma = ws.cell(counter_rows, 9).value
                while " " == address_without_comma[len(address_without_comma) - 1]:
                    address_without_comma = address_without_comma[:-1]

                location = geolocator.geocode(address_without_comma + " Manhattan New York" , country_codes = "US", timeout = 10, bounded=True, viewbox=["40.679807,-74.040816", "40.884657,-73.884947"])

                if location is None:
                    location = geolocator.geocode(search_for_correct_street(address_without_comma, array_of_streets_brooklyn) + " Manhattan New York", country_codes = "US", timeout = 10, bounded=True, viewbox=["40.679807,-74.040816", "40.884657,-73.884947"])

                if not(location is None):
                    ws.cell(counter_rows, 11).value = str(location.latitude) + ', ' + str(location.longitude)
                    print(str(counter_rows) + " Writing Coordinates for : " + ws.cell(counter_rows, 9).value + " " + str(location.latitude) + " " + str(location.longitude))
            counter_rows += 1
    except:
        wb.save("manhattan_found_coordinates_2019.xlsx")
        wb.close()

    wb.save("manhattan_found_coordinates_2019.xlsx")
main()