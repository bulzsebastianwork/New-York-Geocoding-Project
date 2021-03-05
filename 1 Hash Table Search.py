from openpyxl import load_workbook

def main():

    dict_of_addresses = {}

    wb_huge_table = load_workbook("Manhattan_Addresses_Huge_Table.xlsx")
    ws_huge_table = wb_huge_table.active

    Counter_Rows = 2

    while not(ws_huge_table.cell(Counter_Rows, 2).value is None):
        address = str(ws_huge_table.cell(Counter_Rows, 2).value).upper()

        address_split_space = address.split(" ")

        if address_split_space[len(address_split_space) - 1] == "ST":
            address_split_space[len(address_split_space) - 1] = "STREET"
        elif address_split_space[len(address_split_space) - 1] == "AVE":
            address_split_space[len(address_split_space) - 1] = "AVENUE"
        elif address_split_space[len(address_split_space) - 1] == "RD":
            address_split_space[len(address_split_space) - 1] = "ROAD"
        elif address_split_space[len(address_split_space) - 1] == "PL":
            address_split_space[len(address_split_space) - 1] = "PLACE"
        elif address_split_space[len(address_split_space) - 1] == "DR":
            address_split_space[len(address_split_space) - 1] = "DRIVE"
        elif address_split_space[len(address_split_space) - 1] == "CT":
            address_split_space[len(address_split_space) - 1] = "COURT"
        elif address_split_space[len(address_split_space) - 1] == "BLVD":
            address_split_space[len(address_split_space) - 1] = "BOULEVARD"
        elif address_split_space[len(address_split_space) - 1] == "PKWY":
            address_split_space[len(address_split_space) - 1] = "PARKWAY"
        elif address_split_space[len(address_split_space) - 1] == "EXPWY":
            address_split_space[len(address_split_space) - 1] = "EXPRESSWAY"


        dict_of_addresses[" ".join(address_split_space)] = str(ws_huge_table.cell(Counter_Rows, 3).value) + ", " + str(ws_huge_table.cell(Counter_Rows, 4).value)


        Counter_Rows += 1

    wb_addresses_sales = load_workbook("manhattan_2019.xlsx")
    ws_addresses_sales = wb_addresses_sales.active

    array_of_addresses = []
    Counter_Rows = 5
    counter_addresses_found = 0

    while not(ws_addresses_sales.cell(Counter_Rows, 9).value is None):

        ws_addresses_sales.cell(Counter_Rows, 9).value = " ".join(str(ws_addresses_sales.cell(Counter_Rows, 9).value).split())
        if "," in str(ws_addresses_sales.cell(Counter_Rows, 9).value):
            address_split_comma = str(ws_addresses_sales.cell(Counter_Rows, 9).value.upper()).split(",")
            array_of_addresses.append(address_split_comma[0])
        else:
            array_of_addresses.append(str(ws_addresses_sales.cell(Counter_Rows, 9).value).upper())

        if ws_addresses_sales.cell(Counter_Rows, 11).value is None:

            if "APT" in array_of_addresses[Counter_Rows - 5]:
                array_of_addresses[Counter_Rows - 5] = array_of_addresses[Counter_Rows - 5][:array_of_addresses[Counter_Rows - 5].find("APT") - 1]

            address_split_space = array_of_addresses[Counter_Rows - 5].split(" ")
            if "ST." == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "STREET"
            elif "ST" == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "STREET"
            elif "AVE." == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "AVENUE"
            elif "AVE" == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "AVENUE"
            elif "PL." == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "PLACE"
            elif "PL" == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "PLACE"
            elif "BLVD." == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "BOULEVARD"
            elif "BLVD" == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "BOULEVARD"
            elif "PKWY." == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "PARKWAY"
            elif "PKWY" == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "PARKWAY"
            elif "EXPWY." == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "EXPRESSWAY"
            elif "EXPWY" == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "EXPRESSWAY"
            elif "RD." == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "ROAD"
            elif "RD" == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "ROAD"
            elif "DR." == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "DRIVE"
            elif "DR" == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "DRIVE"
            elif "CT." == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "COURT"
            elif "CT" == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "COURT"
            elif "ALY." == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "ALLEY"
            elif "ALY" == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "ALLEY"
            elif "LN." == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "LANE"
            elif "LN" == address_split_space[len(address_split_space) - 1]:
                address_split_space[len(address_split_space) - 1] = "LANE"



            array_of_addresses[Counter_Rows - 5] = " ".join(address_split_space)

            if array_of_addresses[Counter_Rows - 5] in dict_of_addresses.keys():
                counter_addresses_found += 1
                ws_addresses_sales.cell(Counter_Rows, 11).value = dict_of_addresses[array_of_addresses[Counter_Rows - 5]]
                print(str(counter_addresses_found) + ". Coordinates for " + array_of_addresses[Counter_Rows - 5] + " are: " + dict_of_addresses[array_of_addresses[Counter_Rows - 5]])
            else:
                if "1TH" in array_of_addresses[Counter_Rows - 5]:
                    array_of_addresses[Counter_Rows - 5] = array_of_addresses[Counter_Rows - 5].replace("1TH", "1")
                elif "1ST" in array_of_addresses[Counter_Rows - 5]:
                    array_of_addresses[Counter_Rows - 5] = array_of_addresses[Counter_Rows - 5].replace("1ST", "1")
                elif "2ND" in array_of_addresses[Counter_Rows - 5]:
                    array_of_addresses[Counter_Rows - 5] = array_of_addresses[Counter_Rows - 5].replace("2ND", "2")
                elif "3RD" in array_of_addresses[Counter_Rows - 5]:
                    array_of_addresses[Counter_Rows - 5] = array_of_addresses[Counter_Rows - 5].replace("3RD", "3")
                elif "4TH" in array_of_addresses[Counter_Rows - 5]:
                    array_of_addresses[Counter_Rows - 5] = array_of_addresses[Counter_Rows - 5].replace("4TH", "4")
                elif "5TH" in array_of_addresses[Counter_Rows - 5]:
                    array_of_addresses[Counter_Rows - 5] = array_of_addresses[Counter_Rows - 5].replace("5TH", "5")
                elif "6TH" in array_of_addresses[Counter_Rows - 5]:
                    array_of_addresses[Counter_Rows - 5] = array_of_addresses[Counter_Rows - 5].replace("6TH", "6")
                elif "7TH" in array_of_addresses[Counter_Rows - 5]:
                    array_of_addresses[Counter_Rows - 5] = array_of_addresses[Counter_Rows - 5].replace("7TH", "7")
                elif "8TH" in array_of_addresses[Counter_Rows - 5]:
                    array_of_addresses[Counter_Rows - 5] = array_of_addresses[Counter_Rows - 5].replace("8TH", "8")
                elif "9TH" in array_of_addresses[Counter_Rows - 5]:
                    array_of_addresses[Counter_Rows - 5] = array_of_addresses[Counter_Rows - 5].replace("9TH", "9")
                if array_of_addresses[Counter_Rows - 5] in dict_of_addresses.keys():
                    counter_addresses_found += 1
                    ws_addresses_sales.cell(Counter_Rows, 11).value = dict_of_addresses[array_of_addresses[Counter_Rows - 5]]
                    print(str(counter_addresses_found) + ". Coordinates for " + array_of_addresses[Counter_Rows - 5] + " are: " + dict_of_addresses[array_of_addresses[Counter_Rows - 5]])

        Counter_Rows += 1


    wb_addresses_sales.save("manhattan_hash_2019.xlsx")
    wb_addresses_sales.close()

main()